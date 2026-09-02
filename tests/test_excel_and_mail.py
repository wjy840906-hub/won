from pathlib import Path

import pytest
from openpyxl import load_workbook

from kind_managed.config import MailConfig
from kind_managed.excel_writer import COLUMNS, write_excel
from kind_managed.mailer import MailError, build_message, send_message

ROWS = [
    {
        "no": "1",
        "market": "유가증권",
        "code": "005930",
        "name": "테스트전자",
        "biz_no": "104-81-18820",
        "corp_reg_no": "110111-0012345",
        "ceo_name": "홍길동",
        "reason": "감사의견 거절",
        "designated_on": "2026-09-02",
        "note": "",
    },
    {
        "no": "2",
        "market": "코스닥",
        "code": "012345",
        "name": "테스트소재",
        "biz_no": "",
        "corp_reg_no": "",
        "ceo_name": "",
        "reason": "자본잠식률 50% 이상",
        "designated_on": "2026-03-20",
        "note": "DART 고유번호 미매칭",
    },
]


def test_write_excel_creates_expected_layout(tmp_path):
    result = write_excel(ROWS, tmp_path / "관리종목.xlsx", as_of="2026-09-02")

    assert result.row_count == 2
    assert result.path.exists()

    sheet = load_workbook(result.path).active
    assert sheet.title == "관리종목_20260902"
    assert "2026-09-02" in sheet.cell(row=1, column=1).value

    headers = [sheet.cell(row=4, column=i).value for i in range(1, len(COLUMNS) + 1)]
    assert headers == [column[0] for column in COLUMNS]

    assert sheet.cell(row=5, column=4).value == "테스트전자"
    assert sheet.cell(row=5, column=5).value == "104-81-18820"
    assert sheet.cell(row=6, column=10).value == "DART 고유번호 미매칭"
    assert sheet.freeze_panes == "A5"
    assert sheet.auto_filter.ref == "A4:J6"


def test_leading_zero_codes_survive_as_text(tmp_path):
    rows = [dict(ROWS[0], code="005930", biz_no="012-34-56789")]
    result = write_excel(rows, tmp_path / "t.xlsx", as_of="2026-09-02")

    sheet = load_workbook(result.path).active
    assert sheet.cell(row=5, column=3).value == "005930"
    assert sheet.cell(row=5, column=3).number_format == "@"
    assert sheet.cell(row=5, column=5).value == "012-34-56789"


def test_write_excel_with_no_rows(tmp_path):
    result = write_excel([], tmp_path / "empty.xlsx", as_of="2026-09-02")

    sheet = load_workbook(result.path).active
    assert result.row_count == 0
    assert sheet.auto_filter.ref is None
    assert "총 0종목" in sheet.cell(row=1, column=1).value


def _mail_config(**overrides):
    base = dict(
        host="smtp.example.com",
        port=587,
        user="bot@example.com",
        password="secret",
        sender="bot@example.com",
        to=["wonjiyun@hanafn.com"],
    )
    base.update(overrides)
    return MailConfig(**base)


def test_build_message_attaches_xlsx(tmp_path):
    attachment = tmp_path / "관리종목_20260902.xlsx"
    write_excel(ROWS, attachment, as_of="2026-09-02")

    message = build_message(
        _mail_config(cc=["team@example.com"]),
        subject="[관리종목] 2026-09-02",
        body_text="본문",
        body_html="<p>본문</p>",
        attachments=[attachment],
    )

    assert message["To"] == "wonjiyun@hanafn.com"
    assert message["Cc"] == "team@example.com"
    parts = list(message.iter_attachments())
    assert len(parts) == 1
    assert parts[0].get_filename() == "관리종목_20260902.xlsx"
    assert parts[0].get_content_type() == (
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )


def test_build_message_rejects_missing_attachment(tmp_path):
    with pytest.raises(MailError, match="첨부할 파일이 없습니다"):
        build_message(_mail_config(), "제목", "본문", attachments=[tmp_path / "nope.xlsx"])


def test_recipients_merge_to_and_cc_without_duplicates():
    config = _mail_config(cc=["wonjiyun@hanafn.com", "team@example.com"])
    assert config.recipients == ["wonjiyun@hanafn.com", "team@example.com"]


def test_validate_reports_missing_host_and_conflicting_tls():
    problems = _mail_config(host="", use_ssl=True, use_starttls=True).validate()
    assert any("SMTP_HOST" in p for p in problems)
    assert any("동시에" in p for p in problems)


def test_send_message_uses_starttls_and_login(monkeypatch):
    calls = []

    class _FakeSMTP:
        def __init__(self, host, port, timeout=None):
            calls.append(("connect", host, port))

        def __enter__(self):
            return self

        def __exit__(self, *args):
            calls.append(("quit",))
            return False

        def ehlo(self):
            calls.append(("ehlo",))

        def starttls(self):
            calls.append(("starttls",))

        def login(self, user, password):
            calls.append(("login", user))

        def send_message(self, message, from_addr=None, to_addrs=None):
            calls.append(("send", from_addr, tuple(to_addrs)))

    monkeypatch.setattr("smtplib.SMTP", _FakeSMTP)
    config = _mail_config()
    send_message(config, build_message(config, "제목", "본문"))

    assert ("connect", "smtp.example.com", 587) in calls
    assert ("starttls",) in calls
    assert ("login", "bot@example.com") in calls
    assert ("send", "bot@example.com", ("wonjiyun@hanafn.com",)) in calls


def test_send_message_refuses_incomplete_config():
    config = _mail_config(host="")
    with pytest.raises(MailError, match="메일 설정이 올바르지 않습니다"):
        send_message(config, build_message(_mail_config(), "제목", "본문"))
