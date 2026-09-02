from pathlib import Path

import pytest
from openpyxl import load_workbook

from kind_managed.config import AppConfig, MailConfig
from kind_managed.kind_client import ManagedStock
from kind_managed.pipeline import build_mail_bodies, collect_rows, run

STOCKS = [
    ManagedStock(name="테스트전자", reason="감사의견 거절", designated_on="2026-09-02",
                 code="005930", market="유가증권"),
    ManagedStock(name="테스트소재", reason="자본잠식률 50% 이상", designated_on="2026-03-20",
                 code="012345", market="코스닥"),
]


class _FakeDart:
    """005930 만 조회에 성공하는 DART 대역."""

    def lookup(self, stock_code="", name=""):
        if stock_code == "005930":
            from kind_managed.dart_client import CompanyInfo

            return (
                CompanyInfo(
                    corp_code="00126380",
                    corp_name="테스트전자",
                    biz_no="104-81-18820",
                    corp_reg_no="110111-0012345",
                    ceo_name="홍길동",
                ),
                "",
            )
        return None, "DART 고유번호 미매칭"


def test_collect_rows_merges_biz_no():
    rows = collect_rows(STOCKS, _FakeDart(), as_of="2026-09-02")

    assert [row["no"] for row in rows] == ["1", "2"]
    assert rows[0]["biz_no"] == "104-81-18820"
    assert rows[0]["ceo_name"] == "홍길동"
    assert rows[0]["note"] == ""
    assert rows[1]["biz_no"] == ""
    assert rows[1]["note"] == "DART 고유번호 미매칭"


def test_collect_rows_without_dart_key_notes_reason():
    rows = collect_rows(STOCKS, None, as_of="2026-09-02")
    assert all(row["biz_no"] == "" for row in rows)
    assert rows[0]["note"] == "DART_API_KEY 미설정"


def test_build_mail_bodies_highlights_today():
    rows = collect_rows(STOCKS, _FakeDart(), as_of="2026-09-02")
    text, body_html = build_mail_bodies(rows, "2026-09-02", "관리종목_20260902.xlsx")

    assert "관리종목 수: 2종목" in text
    assert "당일(2026-09-02) 신규 지정: 1종목" in text
    assert "테스트전자" in text and "테스트소재" not in text.split("[당일 신규 지정]")[1]
    assert "당일 신규 지정" in body_html
    assert "104-81-18820" in body_html


def test_build_mail_bodies_when_nothing_new_today():
    rows = collect_rows(STOCKS, _FakeDart(), as_of="2026-01-01")
    _text, body_html = build_mail_bodies(rows, "2026-01-01", "f.xlsx")
    assert "당일 신규 지정 종목은 없습니다" in body_html


def test_run_writes_excel_and_sends_mail(tmp_path, monkeypatch):
    sent = {}

    def _fake_send(config, message):
        sent["subject"] = message["Subject"]
        sent["to"] = message["To"]
        sent["attachments"] = [part.get_filename() for part in message.iter_attachments()]

    monkeypatch.setattr("kind_managed.pipeline.send_message", _fake_send)
    monkeypatch.setattr("kind_managed.pipeline.now_kst", lambda: __import__("datetime").datetime(2026, 9, 2))

    app_config = AppConfig(dart_api_key="", out_dir=str(tmp_path), cache_dir=str(tmp_path))
    mail_config = MailConfig(host="smtp.example.com", sender="bot@example.com",
                             to=["wonjiyun@hanafn.com"])

    result = run(app_config, mail_config, send_mail=True, stocks=STOCKS)

    assert result.total == 2
    assert result.mail_sent is True
    assert result.excel_path == tmp_path / "관리종목_20260902.xlsx"
    assert result.excel_path.exists()
    assert sent["subject"] == "[관리종목] 2026-09-02 기준 관리종목 현황 (2종목)"
    assert sent["to"] == "wonjiyun@hanafn.com"
    assert sent["attachments"] == ["관리종목_20260902.xlsx"]

    sheet = load_workbook(result.excel_path).active
    assert sheet.cell(row=5, column=4).value == "테스트전자"


def test_run_can_skip_mail(tmp_path, monkeypatch):
    def _boom(*args, **kwargs):
        raise AssertionError("--no-email 인데 메일을 보내려 했습니다")

    monkeypatch.setattr("kind_managed.pipeline.send_message", _boom)
    app_config = AppConfig(dart_api_key="", out_dir=str(tmp_path), cache_dir=str(tmp_path))

    result = run(app_config, MailConfig(), send_mail=False, stocks=STOCKS)

    assert result.mail_sent is False
    assert result.excel_path.exists()


# ------------------------------------------------------------ 수집 시작일 필터

from kind_managed.config import normalize_from_date
from kind_managed.pipeline import filter_by_from_date

WIDE_STOCKS = [
    ManagedStock(name="구건", reason="감사의견 거절", designated_on="2023-03-23", market="유가증권"),
    ManagedStock(name="칠월건", reason="자본잠식", designated_on="2026-07-31", market="코스닥"),
    ManagedStock(name="팔월첫날", reason="시가총액 미달", designated_on="2026-08-01", market="유가증권"),
    ManagedStock(name="팔월건", reason="시가총액 미달", designated_on="2026-08-13", market="유가증권"),
    ManagedStock(name="구월건", reason="실질심사", designated_on="2026-09-02", market="코스닥"),
]


@pytest.mark.parametrize(
    "raw,expected",
    [
        ("2026-08-01", "2026-08-01"),
        ("2026-08", "2026-08-01"),
        ("2026.8", "2026-08-01"),
        ("20260801", "2026-08-01"),
        ("202608", "2026-08-01"),
        ("", ""),
    ],
)
def test_normalize_from_date(raw, expected):
    assert normalize_from_date(raw) == expected


def test_normalize_from_date_rejects_garbage():
    with pytest.raises(ValueError, match="FROM_DATE"):
        normalize_from_date("작년쯤")


def test_filter_keeps_august_onward():
    kept, dropped = filter_by_from_date(WIDE_STOCKS, "2026-08-01")

    assert [stock.name for stock in kept] == ["팔월첫날", "팔월건", "구월건"]
    assert dropped == 2


def test_filter_boundary_is_inclusive():
    kept, _ = filter_by_from_date(WIDE_STOCKS, "2026-08-01")
    assert "팔월첫날" in [stock.name for stock in kept]


def test_filter_without_from_date_keeps_everything():
    kept, dropped = filter_by_from_date(WIDE_STOCKS, "")
    assert len(kept) == len(WIDE_STOCKS)
    assert dropped == 0


def test_filter_keeps_rows_with_unparsed_date():
    """지정일을 읽지 못한 행은 누락을 막기 위해 남긴다."""
    stocks = [*WIDE_STOCKS, ManagedStock(name="날짜없음", reason="?", designated_on="")]
    kept, _ = filter_by_from_date(stocks, "2026-08-01")
    assert "날짜없음" in [stock.name for stock in kept]


def test_run_applies_from_date_and_labels_period(tmp_path, monkeypatch):
    monkeypatch.setattr(
        "kind_managed.pipeline.now_kst",
        lambda: __import__("datetime").datetime(2026, 9, 2),
    )
    app_config = AppConfig(
        dart_api_key="", out_dir=str(tmp_path), cache_dir=str(tmp_path), from_date="2026-08-01"
    )

    result = run(app_config, MailConfig(), send_mail=False, stocks=WIDE_STOCKS)

    assert result.total == 3
    sheet = load_workbook(result.excel_path).active
    assert "2026-08-01 이후 지정분" in sheet.cell(row=1, column=1).value


def test_run_subject_shows_period(tmp_path, monkeypatch):
    sent = {}
    monkeypatch.setattr(
        "kind_managed.pipeline.send_message",
        lambda config, message: sent.update(subject=message["Subject"]),
    )
    monkeypatch.setattr(
        "kind_managed.pipeline.now_kst",
        lambda: __import__("datetime").datetime(2026, 9, 2),
    )
    app_config = AppConfig(
        dart_api_key="", out_dir=str(tmp_path), cache_dir=str(tmp_path), from_date="2026-08-01"
    )
    mail_config = MailConfig(host="smtp.example.com", sender="a@b.com", to=["c@d.com"])

    run(app_config, mail_config, send_mail=True, stocks=WIDE_STOCKS)

    assert sent["subject"] == "[관리종목] 2026-09-02 기준 2026-08-01~ 관리종목 현황 (3종목)"


def test_collect_rows_fills_stock_code_from_dart():
    """KIND 관리종목 표에는 종목코드가 없으므로 DART 조회 결과로 채운다."""

    class _NameMatchingDart:
        def lookup(self, stock_code="", name=""):
            from kind_managed.dart_client import CompanyInfo

            assert stock_code == ""  # KIND 가 코드를 주지 않는다
            return CompanyInfo(corp_code="00126380", stock_code="5930",
                               biz_no="104-81-18820"), ""

    stock = ManagedStock(name="테스트전자", reason="감사의견 거절",
                         designated_on="2026-08-13", market="유가증권")
    rows = collect_rows([stock], _NameMatchingDart(), as_of="2026-09-02")

    assert rows[0]["code"] == "005930"  # 앞자리 0 보정
    assert rows[0]["biz_no"] == "104-81-18820"
