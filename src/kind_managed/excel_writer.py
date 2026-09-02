"""관리종목 결과를 엑셀(.xlsx) 파일로 저장한다."""

from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(color="FFFFFF", bold=True, size=10)
BODY_FONT = Font(size=10)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# (헤더, 필드명, 너비, 정렬)
COLUMNS: list[tuple[str, str, int, str]] = [
    ("번호", "no", 6, "center"),
    ("시장구분", "market", 10, "center"),
    ("종목코드", "code", 10, "center"),
    ("종목명", "name", 28, "left"),
    ("사업자등록번호", "biz_no", 16, "center"),
    ("법인등록번호", "corp_reg_no", 16, "center"),
    ("대표자", "ceo_name", 12, "center"),
    ("지정사유", "reason", 46, "left"),
    ("지정일", "designated_on", 12, "center"),
    ("비고", "note", 22, "left"),
]


@dataclass
class ExcelResult:
    """엑셀 생성 결과."""

    path: Path
    row_count: int


def _sheet_title(as_of: str) -> str:
    return f"관리종목_{as_of.replace('-', '')}"[:31]


def write_excel(
    rows: list[dict[str, str]],
    out_path: str | Path,
    as_of: str,
    period: str = "",
    source_note: str = "출처: 한국거래소 KIND(kind.krx.co.kr) / 사업자등록번호: 금융감독원 DART",
) -> ExcelResult:
    """관리종목 목록을 서식이 적용된 엑셀 파일로 저장한다."""
    path = Path(out_path)
    path.parent.mkdir(parents=True, exist_ok=True)

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = _sheet_title(as_of)

    scope = f", {period}" if period else ""
    title = f"관리종목 현황 ({as_of} 기준{scope}, 총 {len(rows)}종목)"
    sheet.cell(row=1, column=1, value=title).font = Font(size=13, bold=True)
    sheet.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    sheet.cell(row=2, column=1, value=source_note).font = Font(size=9, color="808080")
    sheet.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))

    header_row = 4
    for index, (header, _field, width, _align) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=header_row, column=index, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.row_dimensions[header_row].height = 22

    for offset, row in enumerate(rows):
        excel_row = header_row + 1 + offset
        for index, (_header, field_name, _width, align) in enumerate(COLUMNS, start=1):
            value = "" if row.get(field_name) is None else str(row.get(field_name, ""))
            cell = sheet.cell(row=excel_row, column=index, value=value)
            cell.font = BODY_FONT
            cell.border = BORDER
            cell.alignment = Alignment(
                horizontal=align, vertical="center", wrap_text=(field_name == "reason")
            )
            if field_name in {"code", "biz_no", "corp_reg_no"}:
                # 앞자리 0 이 사라지지 않도록 텍스트 서식 고정
                cell.number_format = "@"

    last_row = header_row + len(rows)
    if rows:
        sheet.auto_filter.ref = f"A{header_row}:{get_column_letter(len(COLUMNS))}{last_row}"
    sheet.freeze_panes = sheet.cell(row=header_row + 1, column=1)

    workbook.save(path)
    return ExcelResult(path=path, row_count=len(rows))
